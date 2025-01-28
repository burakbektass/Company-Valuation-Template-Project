import streamlit as st
from datetime import date
import pandas as pd
from PIL import Image
import yahoo_fin.stock_info as si
from yahoo_fin import options
from openpyxl import load_workbook
import webbrowser
from config import (tr_companies, sheet_names, excel_names, betas, 
                   no_of_common_stock, tickers, sectors, sectors_string)


st.set_page_config(page_title="Company Valuation Template", page_icon=":bar_chart:" , layout="wide")



# local css kullanarak streamlitten gelen gereksiz yazılar kaldırıldı
def local_css(file_name):
    with open(file_name) as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
local_css("style/style.css")


with st.container():
    
    st.title("Welcome to Company Valuation")
    st.write("You can check our estimated stock prices.")
    # image = Image.open('signin-icon.jpeg')
    # st.image(image,use_column_width='auto')
    st.write("---")


#DCF METHOD
def dcf(option,risk_free_rate,market_rate_of_return):
    try:
        book = load_workbook(excel_names[option])
    except FileNotFoundError:
        st.error(f"Excel dosyası bulunamadı: {excel_names[option]}")
        st.info("Lütfen Excel dosyalarının 'financials' klasöründe olduğundan emin olun.")
        st.stop()
    
    financial_sheets = book[sheet_names[option]]
    ticker = tickers[option]
    perpetual_growth = 0.025
    current_price = si.get_live_price(ticker)
    beta = betas[option]
    shares_out = no_of_common_stock[option]
    total_wacc_values = 0
    number_of_wacc_year = 0
    

    # WACC calculation / wd / rd / we / re  2013
    total_debt_2013 = financial_sheets['D57'].value * 1000
    total_equity_2013= financial_sheets['D52'].value * 1000
    if(total_debt_2013 != 0):
        constant = 0
       
        interest_expense_2013 = abs(financial_sheets['P27'].value * 1000)

        total_2013 = total_debt_2013 + total_equity_2013
        wd_2013 = total_debt_2013 /total_2013
        rd_2013 = interest_expense_2013 / total_debt_2013
        we_2013 = total_equity_2013 / (total_equity_2013 + total_debt_2013)
        re = risk_free_rate + (beta * float((market_rate_of_return - risk_free_rate)))
        wacc_2013 = (wd_2013 * rd_2013 * (1 - constant)) + (we_2013 * re)

        total_wacc_values += wacc_2013
        number_of_wacc_year += 1
    


    # WACC calculation / wd / rd / we / re  2014
    total_debt_2014 = financial_sheets['E57'].value * 1000
    total_equity_2014= financial_sheets['E52'].value * 1000
    if(total_debt_2014 !=0):
        constant = 0
        
        interest_expense_2014 = abs(financial_sheets['Q27'].value * 1000)

        total_2014 = total_debt_2014 + total_equity_2014
        wd_2014 = total_debt_2014 /total_2014
        rd_2014 = interest_expense_2014 / total_debt_2014
        we_2014 = total_equity_2014 / (total_equity_2014 + total_debt_2014)
        re = risk_free_rate + (beta * (market_rate_of_return - risk_free_rate))
        wacc_2014 = (wd_2014 * rd_2014 * (1 - constant)) + (we_2014 * re)

        total_wacc_values += wacc_2014
        number_of_wacc_year += 1

    # WACC calculation / wd / rd / we / re  2015
    total_debt_2015 = financial_sheets['F57'].value * 1000
    total_equity_2015= financial_sheets['F52'].value * 1000
    if(total_debt_2015 !=0):
        constant = 0
        
        interest_expense_2015 = abs(financial_sheets['R27'].value * 1000)

        total_2015 = total_debt_2015 + total_equity_2015
        wd_2015 = total_debt_2015 /total_2015
        rd_2015 = interest_expense_2015 / total_debt_2015
        we_2015 = total_equity_2015 / (total_equity_2015 + total_debt_2015)
        re = risk_free_rate + (beta * (market_rate_of_return - risk_free_rate))
        wacc_2015 = (wd_2015 * rd_2015 * (1 - constant)) + (we_2015 * re)

        total_wacc_values += wacc_2015
        number_of_wacc_year += 1

    # WACC calculation / wd / rd / we / re  2016
    total_debt_2016 = financial_sheets['G57'].value * 1000
    total_equity_2016= financial_sheets['G52'].value * 1000
    if(total_debt_2016 !=0):
        constant = 0
        
        interest_expense_2016 = abs(financial_sheets['S27'].value * 1000)

        total_2016 = total_debt_2016 + total_equity_2016
        wd_2016 = total_debt_2016 /total_2016
        rd_2016 = interest_expense_2016 / total_debt_2016
        we_2016 = total_equity_2016 / (total_equity_2016 + total_debt_2016)
        re = risk_free_rate + (beta * (market_rate_of_return - risk_free_rate))
        wacc_2016 = (wd_2016 * rd_2016 * (1 - constant)) + (we_2016 * re)
        total_wacc_values += wacc_2016
        number_of_wacc_year += 1

    # WACC calculation / wd / rd / we / re  2017
    total_debt_2017 = financial_sheets['H57'].value * 1000
    total_equity_2017 = financial_sheets['H52'].value * 1000
    if(total_debt_2017 !=0):
        constant = 0
        
        interest_expense_2017 = abs(financial_sheets['T27'].value * 1000)

        total_2017 = total_debt_2017 + total_equity_2017
        wd_2017 = total_debt_2017 /total_2017
        rd_2017 = interest_expense_2017 / total_debt_2017
        we_2017 = total_equity_2017/ (total_equity_2017 + total_debt_2017)
        re = risk_free_rate + (beta * (market_rate_of_return - risk_free_rate))
        wacc_2017= (wd_2017* rd_2017* (1 - constant)) + (we_2017 * re)

        total_wacc_values += wacc_2017
        number_of_wacc_year += 1


    # WACC calculation / wd / rd / we / re  2018
    total_debt_2018 = financial_sheets['I57'].value * 1000
    total_equity_2018 = financial_sheets['I52'].value * 1000
    if(total_debt_2018 !=0):
        constant = 0
        interest_expense_2018 = abs(financial_sheets['U27'].value * 1000)

        total_2018 = total_debt_2018 + total_equity_2018
        wd_2018 = total_debt_2018 /total_2018
        rd_2018 = interest_expense_2018 / total_debt_2018
        we_2018 = total_equity_2018/ (total_equity_2018 + total_debt_2018)
        re = risk_free_rate + (beta * (market_rate_of_return - risk_free_rate))
        wacc_2018= (wd_2018* rd_2018* (1 - constant)) + (we_2018 * re)

        total_wacc_values += wacc_2018
        number_of_wacc_year += 1


    # WACC calculation / wd / rd / we / re  2019
    total_debt_2019 = financial_sheets['J57'].value * 1000
    total_equity_2019 = financial_sheets['J52'].value * 1000
    if(total_debt_2019 !=0):
        constant = 0
        
        interest_expense_2019 = abs(financial_sheets['V27'].value * 1000)

        total_2019 = total_debt_2019 + total_equity_2019
        wd_2019 = total_debt_2019 /total_2019
        rd_2019 = interest_expense_2019 / total_debt_2019
        we_2019 = total_equity_2019/ (total_equity_2019 + total_debt_2019)
        re = risk_free_rate + (beta * (market_rate_of_return - risk_free_rate))
        wacc_2019= (wd_2019* rd_2019* (1 - constant)) + (we_2019 * re)

        total_wacc_values += wacc_2019
        number_of_wacc_year += 1

    # WACC calculation / wd / rd / we / re  2020
    total_debt_2020 = financial_sheets['K57'].value * 1000
    total_equity_2020 = financial_sheets['K52'].value * 1000
    if(total_debt_2020 !=0):
        constant = 0
        
        interest_expense_2020 = abs(financial_sheets['W27'].value * 1000)

        total_2020 = total_debt_2020 + total_equity_2020
        wd_2020 = total_debt_2020 /total_2020
        rd_2020 = interest_expense_2020 / total_debt_2020
        we_2020 = total_equity_2020/ (total_equity_2020 + total_debt_2020)
        re = risk_free_rate + (beta * (market_rate_of_return - risk_free_rate))
        wacc_2020= (wd_2020* rd_2020* (1 - constant)) + (we_2020 * re)

        total_wacc_values += wacc_2020
        number_of_wacc_year += 1



    # WACC calculation / wd / rd / we / re  2021
    total_debt_2021 = financial_sheets['L57'].value * 1000
    total_equity_2021 = financial_sheets['L52'].value * 1000
    if(total_debt_2021 !=0):
        constant = 0
        interest_expense_2021 = abs(financial_sheets['X27'].value * 1000)

        total_2021 = total_debt_2021 + total_equity_2021
        wd_2021 = total_debt_2021 /total_2021
        rd_2021 = interest_expense_2021 / total_debt_2021
        we_2021 = total_equity_2021/ (total_equity_2021 + total_debt_2021)
        re = risk_free_rate + (beta * (market_rate_of_return - risk_free_rate))
        wacc_2021= (wd_2021* rd_2021* (1 - constant)) + (we_2021 * re)

        total_wacc_values += wacc_2021
        number_of_wacc_year += 1

    

    total_equity_rate_2014 = (total_equity_2014 - total_equity_2013) / total_equity_2014
    total_equity_rate_2015 = (total_equity_2015 - total_equity_2014) / total_equity_2015
    total_equity_rate_2016 = (total_equity_2016 - total_equity_2015) / total_equity_2016
    total_equity_rate_2017 = (total_equity_2017 - total_equity_2016) / total_equity_2017
    total_equity_rate_2018 = (total_equity_2018 - total_equity_2017) / total_equity_2018
    total_equity_rate_2019 = (total_equity_2019 - total_equity_2018) / total_equity_2019
    total_equity_rate_2020 = (total_equity_2020 - total_equity_2019) / total_equity_2020
    total_equity_rate_2021 = (total_equity_2021 - total_equity_2020) / total_equity_2021

    average_total_equity_rate = (total_equity_rate_2014 + total_equity_rate_2015 + total_equity_rate_2016 +
                                 total_equity_rate_2017 + total_equity_rate_2018 + total_equity_rate_2019 +
                                 total_equity_rate_2020 + total_equity_rate_2021) / 8

    total_equity_2022 = total_equity_2021 * (1 + average_total_equity_rate)
    total_equity_2023 = total_equity_2022 * (1 + average_total_equity_rate)
    total_equity_2024 = total_equity_2023 * (1 + average_total_equity_rate)
    total_equity_2025 = total_equity_2024 * (1 + average_total_equity_rate)
    

    #Required Return
    required_return = total_wacc_values / number_of_wacc_year
    #Revenues
    revenue_2013 = financial_sheets['P13'].value * 1000
    revenue_2014 = financial_sheets['Q13'].value * 1000
    revenue_2015 = financial_sheets['R13'].value * 1000
    revenue_2016 = financial_sheets['S13'].value * 1000
    revenue_2017 = financial_sheets['T13'].value * 1000
    revenue_2018 = financial_sheets['U13'].value * 1000
    revenue_2019 = financial_sheets['V13'].value * 1000
    revenue_2020 = financial_sheets['W13'].value * 1000
    revenue_2021 = financial_sheets['X13'].value * 1000

    # Net Incomes 
    net_income_2013 = financial_sheets['P31'].value * 1000
    net_income_2014 = financial_sheets['Q31'].value * 1000
    net_income_2015 = financial_sheets['R31'].value * 1000
    net_income_2016 = financial_sheets['S31'].value * 1000
    net_income_2017 = financial_sheets['T31'].value * 1000
    net_income_2018 = financial_sheets['U31'].value * 1000
    net_income_2019 = financial_sheets['V31'].value * 1000
    net_income_2020 = financial_sheets['W31'].value * 1000
    net_income_2021 = financial_sheets['X31'].value * 1000

    #Revenue Growth Rates
    revenue_growth_rate_2014 = (revenue_2014 - revenue_2013) / revenue_2014
    revenue_growth_rate_2015 = (revenue_2015 - revenue_2014) / revenue_2015
    revenue_growth_rate_2016 = (revenue_2016 - revenue_2015) / revenue_2016
    revenue_growth_rate_2017 = (revenue_2017 - revenue_2016) / revenue_2017
    revenue_growth_rate_2018 = (revenue_2018 - revenue_2017) / revenue_2018
    revenue_growth_rate_2019 = (revenue_2019 - revenue_2018) / revenue_2019
    revenue_growth_rate_2020 = (revenue_2020 - revenue_2019) / revenue_2020
    revenue_growth_rate_2021 = (revenue_2021 - revenue_2020) / revenue_2021
    revenue_growth_rate_2022 = ((revenue_growth_rate_2014 + revenue_growth_rate_2015 + 
                                revenue_growth_rate_2016 + revenue_growth_rate_2017 + 
                                revenue_growth_rate_2018 + revenue_growth_rate_2019 + 
                                revenue_growth_rate_2020 + revenue_growth_rate_2021  ) / 8)

    #st.write("Calculated Revenue Growth Rate is ", revenue_growth_rate_2022, " for 2022")
    
    #Net Income Margins
    net_income_margin_ratio_2013 = net_income_2013 / revenue_2013
    net_income_margin_ratio_2014 = net_income_2014 / revenue_2014
    net_income_margin_ratio_2015 = net_income_2015 / revenue_2015
    net_income_margin_ratio_2016 = net_income_2016 / revenue_2016
    net_income_margin_ratio_2017 = net_income_2017 / revenue_2017
    net_income_margin_ratio_2018 = net_income_2018 / revenue_2018
    net_income_margin_ratio_2019 = net_income_2019 / revenue_2019
    net_income_margin_ratio_2020 = net_income_2020 / revenue_2020
    net_income_margin_ratio_2021 = net_income_2021 / revenue_2021
    net_income_margin_ratios_previous = (net_income_margin_ratio_2013 + net_income_margin_ratio_2014 + 
                                         net_income_margin_ratio_2015 + net_income_margin_ratio_2016 + 
                                         net_income_margin_ratio_2017 + net_income_margin_ratio_2018 + 
                                         net_income_margin_ratio_2019 + net_income_margin_ratio_2020 + 
                                         net_income_margin_ratio_2021)

    net_income_margin_ratio_2022 = net_income_margin_ratios_previous / 9
    net_income_margin_ratio_2023 = (net_income_margin_ratios_previous + net_income_margin_ratio_2022) / 10
    net_income_margin_ratio_2024 = (net_income_margin_ratios_previous + net_income_margin_ratio_2022 + 
                                    net_income_margin_ratio_2023) / 11
    net_income_margin_ratio_2025 = (net_income_margin_ratios_previous + 
                                    net_income_margin_ratio_2022 + net_income_margin_ratio_2023 + net_income_margin_ratio_2024 ) / 12

    # Net Income Estimation
    net_income_2022 = net_income_2021 + (net_income_2021 * net_income_margin_ratio_2022 )
    net_income_2023 = net_income_2022 + (net_income_2022 * net_income_margin_ratio_2023)
    net_income_2024 = net_income_2023 + (net_income_2023 * net_income_margin_ratio_2024)
    net_income_2025 = net_income_2024 + (net_income_2024 * net_income_margin_ratio_2025)
    discount_factor_2022 = (1 + required_return)**1
    discount_factor_2023 = (1 + required_return)**2
    discount_factor_2024 = (1 + required_return)**3
    discount_factor_2025 = (1 + required_return)**4
    discount_factor_terminal_value = (1 + required_return)**4  

    # EAT For Free Cash Flow
    eat_2013 = ((financial_sheets['P13'].value + financial_sheets['P16'].value + financial_sheets['P21'].value + financial_sheets['P22'].value) * 1000) * 0.79
    eat_2014 = ((financial_sheets['Q13'].value + financial_sheets['Q16'].value + financial_sheets['Q21'].value + financial_sheets['Q22'].value) * 1000) * 0.79
    eat_2015 = ((financial_sheets['R13'].value + financial_sheets['R16'].value + financial_sheets['R21'].value + financial_sheets['R22'].value) * 1000) * 0.79
    eat_2016 = ((financial_sheets['S13'].value + financial_sheets['S16'].value + financial_sheets['S21'].value + financial_sheets['S22'].value) * 1000) * 0.79
    eat_2017 = ((financial_sheets['T13'].value + financial_sheets['T16'].value + financial_sheets['T21'].value + financial_sheets['T22'].value) * 1000) * 0.79
    eat_2018 = ((financial_sheets['U13'].value + financial_sheets['U16'].value + financial_sheets['U21'].value + financial_sheets['U22'].value) * 1000) * 0.79
    eat_2019 = ((financial_sheets['V13'].value + financial_sheets['V16'].value + financial_sheets['V21'].value + financial_sheets['V22'].value) * 1000) * 0.79
    eat_2020 = ((financial_sheets['W13'].value + financial_sheets['W16'].value + financial_sheets['W21'].value + financial_sheets['W22'].value) * 1000) * 0.79
    eat_2021 = ((financial_sheets['X13'].value + financial_sheets['X16'].value + financial_sheets['X21'].value + financial_sheets['X22'].value) * 1000) * 0.79

    # Depreciation 
    depreciation_2013 = financial_sheets['P55'].value * 1000
    depreciation_2014 = financial_sheets['Q55'].value * 1000
    depreciation_2015 = financial_sheets['R55'].value * 1000
    depreciation_2016 = financial_sheets['S55'].value * 1000
    depreciation_2017 = financial_sheets['T55'].value * 1000
    depreciation_2018 = financial_sheets['U55'].value * 1000
    depreciation_2019 = financial_sheets['V55'].value * 1000
    depreciation_2020 = financial_sheets['W55'].value * 1000
    depreciation_2021 = financial_sheets['X55'].value * 1000

    # Increase in fixed assets
    increased_in_fixed_assets_2013 = (financial_sheets['D21'].value - financial_sheets['C21'].value) * 1000
    increased_in_fixed_assets_2014 = (financial_sheets['E21'].value - financial_sheets['D21'].value) * 1000
    increased_in_fixed_assets_2015 = (financial_sheets['F21'].value - financial_sheets['E21'].value) * 1000
    increased_in_fixed_assets_2016 = (financial_sheets['G21'].value - financial_sheets['F21'].value) * 1000
    increased_in_fixed_assets_2017 = (financial_sheets['H21'].value - financial_sheets['G21'].value) * 1000
    increased_in_fixed_assets_2018 = (financial_sheets['I21'].value - financial_sheets['H21'].value) * 1000
    increased_in_fixed_assets_2019 = (financial_sheets['J21'].value - financial_sheets['I21'].value) * 1000
    increased_in_fixed_assets_2020 = (financial_sheets['K21'].value - financial_sheets['J21'].value) * 1000
    increased_in_fixed_assets_2021 = (financial_sheets['L21'].value - financial_sheets['K21'].value) * 1000

    # Increase in working capital
    increase_in_working_capital_2013 = (financial_sheets['D19'].value - financial_sheets['D36'].value) * 1000 - (financial_sheets['C19'].value - financial_sheets['C36'].value)  * 1000
    increase_in_working_capital_2014 = (financial_sheets['E19'].value - financial_sheets['E36'].value) * 1000 - (financial_sheets['D19'].value - financial_sheets['D36'].value)  * 1000
    increase_in_working_capital_2015 = (financial_sheets['F19'].value - financial_sheets['F36'].value) * 1000 - (financial_sheets['E19'].value - financial_sheets['E36'].value)  * 1000
    increase_in_working_capital_2016 = (financial_sheets['G19'].value - financial_sheets['G36'].value) * 1000 - (financial_sheets['F19'].value - financial_sheets['F36'].value)  * 1000
    increase_in_working_capital_2017 = (financial_sheets['H19'].value - financial_sheets['H36'].value) * 1000 - (financial_sheets['G19'].value - financial_sheets['G36'].value)  * 1000
    increase_in_working_capital_2018 = (financial_sheets['I19'].value - financial_sheets['I36'].value) * 1000 - (financial_sheets['H19'].value - financial_sheets['H36'].value)  * 1000
    increase_in_working_capital_2019 = (financial_sheets['J19'].value - financial_sheets['J36'].value) * 1000 - (financial_sheets['I19'].value - financial_sheets['I36'].value)  * 1000
    increase_in_working_capital_2020 = (financial_sheets['K19'].value - financial_sheets['K36'].value) * 1000 - (financial_sheets['J19'].value - financial_sheets['J36'].value)  * 1000
    increase_in_working_capital_2021 = (financial_sheets['L19'].value - financial_sheets['L36'].value) * 1000 - (financial_sheets['K19'].value - financial_sheets['K36'].value)  * 1000

    # Free cash Flow of previous years
    free_cash_flow_2013 = eat_2013 + depreciation_2013 - increased_in_fixed_assets_2013 - increase_in_working_capital_2013
    free_cash_flow_2014 = eat_2014 + depreciation_2014 - increased_in_fixed_assets_2014 - increase_in_working_capital_2014
    free_cash_flow_2015 = eat_2015 + depreciation_2015 - increased_in_fixed_assets_2015 - increase_in_working_capital_2015
    free_cash_flow_2016 = eat_2016 + depreciation_2016 - increased_in_fixed_assets_2016 - increase_in_working_capital_2016
    free_cash_flow_2017 = eat_2017 + depreciation_2017 - increased_in_fixed_assets_2017 - increase_in_working_capital_2017
    free_cash_flow_2018 = eat_2018 + depreciation_2018 - increased_in_fixed_assets_2018 - increase_in_working_capital_2018
    free_cash_flow_2019 = eat_2019 + depreciation_2019 - increased_in_fixed_assets_2019 - increase_in_working_capital_2019
    free_cash_flow_2020 = eat_2020 + depreciation_2020 - increased_in_fixed_assets_2020 - increase_in_working_capital_2020
    free_cash_flow_2021 = eat_2021 + depreciation_2021 - increased_in_fixed_assets_2021 - increase_in_working_capital_2021

    # CapEx 
    capex_2013 = (financial_sheets['D21'].value - financial_sheets['C21'].value + financial_sheets['P55'].value )*1000
    capex_2014 = (financial_sheets['E21'].value - financial_sheets['D21'].value + financial_sheets['Q55'].value )*1000
    capex_2015 = (financial_sheets['F21'].value - financial_sheets['E21'].value + financial_sheets['R55'].value )*1000
    capex_2016 = (financial_sheets['G21'].value - financial_sheets['F21'].value + financial_sheets['S55'].value )*1000
    capex_2017 = (financial_sheets['H21'].value - financial_sheets['G21'].value + financial_sheets['T55'].value )*1000
    capex_2018 = (financial_sheets['I21'].value - financial_sheets['H21'].value + financial_sheets['U55'].value )*1000
    capex_2019 = (financial_sheets['J21'].value - financial_sheets['I21'].value + financial_sheets['V55'].value )*1000
    capex_2020 = (financial_sheets['K21'].value - financial_sheets['J21'].value + financial_sheets['W55'].value )*1000
    capex_2021 = (financial_sheets['L21'].value - financial_sheets['K21'].value + financial_sheets['X55'].value )*1000

    # Net Borrowing
    net_borrowing_2013 = (financial_sheets['D39'].value - financial_sheets['D34'].value)*1000
    net_borrowing_2014 = (financial_sheets['E39'].value - financial_sheets['E34'].value)*1000
    net_borrowing_2015 = (financial_sheets['F39'].value - financial_sheets['F34'].value)*1000
    net_borrowing_2016 = (financial_sheets['G39'].value - financial_sheets['G34'].value)*1000
    net_borrowing_2017 = (financial_sheets['H39'].value - financial_sheets['H34'].value)*1000
    net_borrowing_2018 = (financial_sheets['I39'].value - financial_sheets['I34'].value)*1000
    net_borrowing_2019 = (financial_sheets['J39'].value - financial_sheets['J34'].value)*1000
    net_borrowing_2020 = (financial_sheets['K39'].value - financial_sheets['K34'].value)*1000
    net_borrowing_2021 = (financial_sheets['L39'].value - financial_sheets['L34'].value)*1000

    #Free Cash Flow to Equity
    fcfe_2013 = net_income_2013 + depreciation_2013 - increase_in_working_capital_2013 - capex_2013 + net_borrowing_2013
    fcfe_2014 = net_income_2014 + depreciation_2014 - increase_in_working_capital_2014 - capex_2014 + net_borrowing_2014
    fcfe_2015 = net_income_2015 + depreciation_2015 - increase_in_working_capital_2015 - capex_2015 + net_borrowing_2015
    fcfe_2016 = net_income_2016 + depreciation_2016 - increase_in_working_capital_2016 - capex_2016 + net_borrowing_2016
    fcfe_2017 = net_income_2017 + depreciation_2017 - increase_in_working_capital_2017 - capex_2017 + net_borrowing_2017
    fcfe_2018 = net_income_2018 + depreciation_2018 - increase_in_working_capital_2018 - capex_2018 + net_borrowing_2018
    fcfe_2019 = net_income_2019 + depreciation_2019 - increase_in_working_capital_2019 - capex_2019 + net_borrowing_2019
    fcfe_2020 = net_income_2020 + depreciation_2020 - increase_in_working_capital_2020 - capex_2020 + net_borrowing_2020
    fcfe_2021 = net_income_2021 + depreciation_2021 - increase_in_working_capital_2021 - capex_2021 + net_borrowing_2021

    # FCF / Net Income
    fcf_over_nı_2013 = fcfe_2013 / net_income_2013
    fcf_over_nı_2014 = fcfe_2014 / net_income_2014
    fcf_over_nı_2015 = fcfe_2015 / net_income_2015
    fcf_over_nı_2016 = fcfe_2016 / net_income_2016
    fcf_over_nı_2017 = fcfe_2017 / net_income_2017
    fcf_over_nı_2018 = fcfe_2018 / net_income_2018
    fcf_over_nı_2019 = fcfe_2019 / net_income_2019
    fcf_over_nı_2020 = fcfe_2020 / net_income_2020
    fcf_over_nı_2021 = fcfe_2021 / net_income_2021
    fcf_over_nı_array = [fcf_over_nı_2013,fcf_over_nı_2014,fcf_over_nı_2015,
                        fcf_over_nı_2016,fcf_over_nı_2017,fcf_over_nı_2018,
                        fcf_over_nı_2019,fcf_over_nı_2020,fcf_over_nı_2021]
    fcf_over_nı_average = (fcf_over_nı_2013 + fcf_over_nı_2014 + fcf_over_nı_2015 + 
                        fcf_over_nı_2016 + fcf_over_nı_2017 + fcf_over_nı_2018 + 
                        fcf_over_nı_2019 + fcf_over_nı_2020 + fcf_over_nı_2021) / 9

    if(abs(fcf_over_nı_average) > max(fcf_over_nı_array)):
        fcf_over_nı_2022 = max(fcf_over_nı_array)
    else:
        fcf_over_nı_2022 = abs(fcf_over_nı_average)
        
    fcf_over_nı_2023 = fcf_over_nı_2022
    fcf_over_nı_2024 = fcf_over_nı_2022
    fcf_over_nı_2025 = fcf_over_nı_2022

    # Free cash Flow estimation 2022, 2023, 2024, 2025
    free_cash_flow_2022 = net_income_2022 + (fcf_over_nı_2022 * net_income_2022)
    free_cash_flow_2023 = net_income_2023 + (fcf_over_nı_2023 * net_income_2023)
    free_cash_flow_2024 = net_income_2024 + (fcf_over_nı_2024 * net_income_2024)
    free_cash_flow_2025 = net_income_2025 + (fcf_over_nı_2025 * net_income_2025)
    free_cash_flow_terminal_value = (free_cash_flow_2025 * (1 + perpetual_growth)) / required_return - perpetual_growth 

    # PV of future cash flow
    pv_of_future_cash_flow_2022 = free_cash_flow_2022 / discount_factor_2022
    pv_of_future_cash_flow_2023 = free_cash_flow_2023 / discount_factor_2023
    pv_of_future_cash_flow_2024 = free_cash_flow_2024 / discount_factor_2024
    pv_of_future_cash_flow_2025 = free_cash_flow_2025 / discount_factor_2025
    pv_of_future_cash_flow_terminal_value = free_cash_flow_terminal_value /discount_factor_terminal_value

    # Residual Incomes
    residual_income_2019 = (financial_sheets['V31'].value * 1000) - (financial_sheets['J52'].value * 1000) * re
    residual_income_2020 = (financial_sheets['W31'].value * 1000) - (financial_sheets['K52'].value * 1000) * re
    residual_income_2021 = (financial_sheets['X31'].value * 1000) - (financial_sheets['L52'].value * 1000) * re
    residual_income_2022 = (net_income_2022 - (re * (financial_sheets['L52'].value * 1000 ))) 
    residual_income_2023 = (net_income_2023 - (re * total_equity_2022 )) 
    residual_income_2024 = (net_income_2024 - (re * total_equity_2023)) 
    residual_income_2025 = (net_income_2025 - (re * total_equity_2024)) 
    residual_income_terminal_value = residual_income_2025 * (1 + 0.025)
    terminal_value_residual = (residual_income_terminal_value / re) - 0.025

    pv_of_residual_income_2022 = residual_income_2022 /(1 + re)**1
    pv_of_residual_income_2023 = residual_income_2023 /(1 + re)**2
    pv_of_residual_income_2024 = residual_income_2024 /(1 + re)**3
    pv_of_residual_income_2025 = residual_income_2025 /(1 + re)**4
    pv_of_residual_income_terminal_value = terminal_value_residual / (1 + re)**4

    total_equity_value = ((pv_of_residual_income_2025 + pv_of_residual_income_2024 
                        + pv_of_residual_income_2023 + pv_of_residual_income_2022) 
                        + total_equity_2021 + pv_of_residual_income_terminal_value)

    
    dcf.price_per_share = abs(total_equity_value / shares_out)
    with st.expander("SEE RESIDUAL INCOME METHOD OUTPUT HERE"):
        st.write("Price per share calculated with Residual Income Method is ",dcf.price_per_share)

    todays_company_value = (pv_of_future_cash_flow_2022 +  pv_of_future_cash_flow_2023 
                     + pv_of_future_cash_flow_2024 + pv_of_future_cash_flow_2025 
                     + pv_of_future_cash_flow_terminal_value)
    
    #Fair Value of Company
    dcf.fair_value_of_company = abs(todays_company_value / shares_out)


    if(dcf.fair_value_of_company < current_price):
        with st.expander("SEE DCF OUTPUT HERE"):
            
            st.write("Fair value of company calculated with DCF: ",dcf.fair_value_of_company)
            st.write("Current stock price of company : ",current_price)
            st.write("OVER PRICED!!")
    else:
        with st.expander("SEE DCF OUTPUT HERE"):
            
            st.write("Fair value of company calculated with DCF: ",dcf.fair_value_of_company)
            st.write("Current stock price of company: ",current_price)
            st.write("UNDER PRICED!!")

    
    
     
    # with st.expander("SEE WEIGHTED VALUE OF DCF AND RESIDUAL INCOME IS HERE"):  
    #     st.write("Weighted price is: ", weighted_price) 
                                   #33333333333333333333333333333333333333333333333333333333333333333333333333333333333    BURASI DEĞİŞMELİ
    
    return dcf.fair_value_of_company,dcf.price_per_share   # weighted_price için eklendi



#Enterprise METHOD
def enterprise_formul(company): 
    try:
        book = load_workbook(excel_names[company])
    except FileNotFoundError:
        st.error(f"Excel dosyası bulunamadı: {excel_names[company]}")
        st.info("Lütfen Excel dosyasının doğru konumda olduğundan emin olun.")
        st.stop()
    financial_sheets = book[sheet_names[company]]
    #2021
    current_price= si.get_live_price(tickers[company])
    enterprise_formul.marketcap = current_price * no_of_common_stock[company]
    enterprise_formul.totaldebt = financial_sheets['L57'].value * 1000
    enterprise_formul.cashandcasheq = (financial_sheets['L13'].value) * 1000
    enterprise_formul.EBITDA = (financial_sheets['X31'].value +
         abs(financial_sheets['X30'].value) +
         abs(financial_sheets['X27'].value) +
         financial_sheets['X55'].value) * 1000
    enterprise_formul.enterprisevalue = enterprise_formul.marketcap +  enterprise_formul.totaldebt -  enterprise_formul.cashandcasheq
    enterprise_formul.valueofcompany = abs(enterprise_formul.enterprisevalue / enterprise_formul.EBITDA)
    
    return enterprise_formul.valueofcompany
# Sidebar
st.sidebar.header('Welcome to Company Valuation Template')

# Companies and constants
option = st.sidebar.selectbox('Companies', ('Google','Apple','Microsoft',
                        'Coca-Cola','PepsiCo','Monster',
                        'Walmart','Dollar General Corporation',
                        'Target','Aksa Enerji','Odaş Elektrik',
                        'Ak Enerji','Migros','Carrefour SA','BIM',
                        'Arcelik','Vestel','Bosch'))
ticker = tickers[option]
calculate = st.sidebar.button('Calculate',key=1,)
st.sidebar.text("")
st.sidebar.write('Or you can upload your own financial sheets.')
st.sidebar.text("")

#***********************************BURASI UPLOAD KISMI*************************
upload_file = st.sidebar.file_uploader('Choose a file to upload', type = 'xlsx')

if upload_file :
    st.sidebar.markdown('---')
    book = load_workbook(upload_file)
    financial_sheets = book[book.sheetnames[0]]
    if book.sheetnames[0] in tr_companies:
        risk_free_rate = 0.17 
        market_rate_of_return = 0.583 
        constant = 0.23
        st.sidebar.write("Turkish company ")
        st.sidebar.write("Risk free rate for Turkey: ", risk_free_rate)
        st.sidebar.write("Market rate of return for Turkey: ", market_rate_of_return)
        st.sidebar.write("Beta: ", betas[book.sheetnames[0]])
        st.sidebar.write("Number of common stocks: ", no_of_common_stock[book.sheetnames[0]])
        st.write("Financial sheet of ",book.sheetnames[0]," Company is used.")
        dcf(book.sheetnames[0],risk_free_rate,market_rate_of_return)
        sector_list = sectors[book.sheetnames[0]]
        total_sector_value = 0
        no_of_sector_company = 0
        total_stock_price_of_sector = 0  #yeni
        total_evm_price = 0

        with st.expander("SEE ENTERPRISE OUTPUT HERE"):
            for company in sector_list:
                enterprise_formul(company)
                total_sector_value += enterprise_formul.valueofcompany
                no_of_sector_company += 1  
                current_price = si.get_live_price(tickers[company]) #yeni
                total_stock_price_of_sector += current_price  #yeni
                average_stock_price_of_sector = total_stock_price_of_sector /no_of_sector_company #yeni 
                sector_average = total_sector_value / no_of_sector_company
                evm_stock_price = ((average_stock_price_of_sector/sector_average)*enterprise_formul.valueofcompany) #son
                total_evm_price += evm_stock_price #son
                average_evm_value = total_evm_price / no_of_sector_company #son
                st.write("Enterprise Ratio of ", company ,": ",enterprise_formul.valueofcompany)
            enterprise_formul(book.sheetnames[0]) # yeni enterprise value değeri için
            evm_stock_price = ((average_stock_price_of_sector/sector_average)*enterprise_formul.valueofcompany) # yeni enterprise value değeri için
            st.write("Enterprise Value for ",book.sheetnames[0] ," is: ",evm_stock_price) # yeni enterprise value değeri için
            st.write("Average stock price of ", sectors_string[option], " is ",average_stock_price_of_sector)  #yeni
            st.write("Average Enterprise Ratio for ", sectors_string[option], " is ",sector_average)

            if(evm_stock_price > average_evm_value):         #enterprise parametresi option değil book.sheetnames[0] oldu
                st.write("Value of ",book.sheetnames[0], " is greater than average.")   #option değil book.sheetnames[0] oldu
                st.write("Overpriced")
            else:
                st.write("Value of ",book.sheetnames[0], " is lower than average.")  # option değil book.sheetnames[0] oldu
                st.write("Underpriced")
        with st.expander("SEE WEIGHTED VALUE OF ALL METHODS OUTPUT HERE"):
            weighted_value = 0.4 * (dcf.fair_value_of_company) + 0.1 * dcf.price_per_share + 0.5 * evm_stock_price ## yeni weighted için
            st.write("Weighted Value", weighted_value) ## yeni weighted için
            st.write("Current stock price ",si.get_live_price(tickers[book.sheetnames[0]]))
            if(weighted_value > si.get_live_price(tickers[book.sheetnames[0]])):
                st.write("Weighted value of ",book.sheetnames[0], " is greater than current stock price.")
                st.write("Decision: Underpriced")
            else:
                st.write("Weighted value of ",book.sheetnames[0], " is lower than current stock price.")
                st.write("Decision: Overpriced")
    else:
        risk_free_rate = 0.029 
        market_rate_of_return = 0.10 
        constant = 0.21
        st.sidebar.write("Turkish company ")
        st.sidebar.write("Risk free rate for Turkey: ", risk_free_rate)
        st.sidebar.write("Market rate of return for Turkey: ", market_rate_of_return)
        st.sidebar.write("Beta: ", betas[book.sheetnames[0]])
        st.sidebar.write("Number of common stocks: ", no_of_common_stock[book.sheetnames[0]])
        st.write("Financial sheet of ",book.sheetnames[0]," Company is used.")
        dcf(book.sheetnames[0],risk_free_rate,market_rate_of_return)
        sector_list = sectors[book.sheetnames[0]] #option değil book.sheetnames[0] oldu
        total_sector_value = 0
        no_of_sector_company = 0
        total_stock_price_of_sector = 0  #yeni
        total_evm_price = 0

        with st.expander("SEE ENTERPRISE OUTPUT HERE"):
            for company in sector_list:
                enterprise_formul(company)
                total_sector_value += enterprise_formul.valueofcompany
                no_of_sector_company += 1
                current_price = si.get_live_price(tickers[company]) #yeni
                total_stock_price_of_sector += current_price    #yeni
                average_stock_price_of_sector = total_stock_price_of_sector /no_of_sector_company #yeni
                sector_average = total_sector_value / no_of_sector_company
                evm_stock_price = ((average_stock_price_of_sector/sector_average)*enterprise_formul.valueofcompany) #son
                total_evm_price += evm_stock_price #son
                average_evm_value = total_evm_price / no_of_sector_company #son
                st.write("Enterprise Ratio of ", company ,": ",enterprise_formul.valueofcompany)
            enterprise_formul(book.sheetnames[0]) # yeni enterprise value değeri için
            evm_stock_price = ((average_stock_price_of_sector/sector_average)*enterprise_formul.valueofcompany) # yeni enterprise value değeri için
            st.write("Enterprise Value for ",book.sheetnames[0] ," is: ",evm_stock_price) # yeni enterprise value değeri için
            st.write("Average stock price of ", sectors_string[book.sheetnames[0]], " is ",average_stock_price_of_sector)  #yeni
            st.write("Average Enterprise Ratio for ", sectors_string[book.sheetnames[0]], " is ",sector_average) #sectors_string[book.sheetnames[0]] oldu

            if(evm_stock_price > average_evm_value):     #enterprise parametresi option değil book.sheetnames[0] oldu
                st.write("Value of ",book.sheetnames[0], " is greater than average.")   #option değil book.sheetnames[0] oldu
                st.write("Overpriced")
            else:
                st.write("Value of ",book.sheetnames[0], " is lower than average.")    #option değil book.sheetnames[0] oldu
                st.write("Underpriced")
        with st.expander("SEE WEIGHTED VALUE OF ALL METHODS OUTPUT HERE"):
            weighted_value = 0.4 * (dcf.fair_value_of_company) + 0.5 * dcf.price_per_share + 0.1 * evm_stock_price ## yeni weighted için
            st.write("Weighted Value", weighted_value) ## yeni weighted için
            st.write("Current stock price ",si.get_live_price(tickers[book.sheetnames[0]]))
            if(weighted_value > si.get_live_price(tickers[book.sheetnames[0]])):
                st.write("Weighted value of ",book.sheetnames[0], " is greater than current stock price.")
                st.write("Decision: Underpriced")
            else:
                st.write("Weighted value of ",book.sheetnames[0], " is lower than current stock price.")
                st.write("Decision: Overpriced")


#***********************************BURASI UPLOAD KISMI*************************
    
if calculate:
    if option in tr_companies:
        risk_free_rate = 0.17 # us 0.029
        market_rate_of_return = 0.583 # us 0.10
        constant = 0.23 # us 0.21
        st.sidebar.write("Turkish company ")
        st.sidebar.write("Risk free rate for Turkey: ", risk_free_rate)
        st.sidebar.write("Market rate of return for Turkey: ", market_rate_of_return)
        st.sidebar.write("Beta: ", betas[option])
        st.sidebar.write("Number of common stocks: ", no_of_common_stock[option])
        book = load_workbook(excel_names[option])
        financial_sheets = book[sheet_names[option]]
        fs_rows = financial_sheets.rows
        st.write("Financial sheet of ",option," Company is used.")
        dcf(option,risk_free_rate,market_rate_of_return)

        sector_list = sectors[option]
        total_sector_value = 0
        no_of_sector_company = 0
        total_stock_price_of_sector = 0  #yeni
        total_evm_price = 0

        with st.expander("SEE ENTERPRISE OUTPUT HERE"):
            for company in sector_list:

                enterprise_formul(company)
                total_sector_value += enterprise_formul.valueofcompany
                no_of_sector_company += 1  
                current_price = si.get_live_price(tickers[company]) #yeni
                total_stock_price_of_sector += current_price           #yeni
                sector_average = total_sector_value / no_of_sector_company
                average_stock_price_of_sector = total_stock_price_of_sector /no_of_sector_company #yeni 
                sector_average = total_sector_value / no_of_sector_company
                evm_stock_price = ((average_stock_price_of_sector/sector_average)*enterprise_formul.valueofcompany) #son
                total_evm_price += evm_stock_price #son
                average_evm_value = total_evm_price / no_of_sector_company #son
                st.write("Enterprise Ratio of ", company ,": ",enterprise_formul.valueofcompany)
            enterprise_formul(option) # yeni enterprise value değeri için
            evm_stock_price = ((average_stock_price_of_sector/sector_average)*enterprise_formul.valueofcompany) # yeni enterprise value değeri için
            st.write("Enterprise Value for ",book.sheetnames[0] ," is: ",evm_stock_price) # yeni enterprise value değeri için
            st.write("Average stock price of ", sectors_string[option], " is ",average_stock_price_of_sector) #yeni
            st.write("Average Enterprise Ratio for ", sectors_string[option], " is ",sector_average)
            if(evm_stock_price > average_evm_value):
                st.write("value of ",option, " is greater than average.")
                st.write("Decision: Overpriced")
            else:
                st.write("value of ",option, " is lower than average.")
                st.write("Decision: Underpriced")
        with st.expander("SEE WEIGHTED VALUE OF ALL METHODS OUTPUT HERE"):
            weighted_value = 0.4 * (dcf.fair_value_of_company) + 0.1 * dcf.price_per_share + 0.5 * evm_stock_price ## yeni weighted için
            st.write("Weighted Value", weighted_value) ## yeni weighted için
            st.write("Current stock price ",si.get_live_price(tickers[option]))
            if(weighted_value > si.get_live_price(tickers[option])):
                st.write("Weighted value of ",option, " is greater than current stock price.")
                st.write("Decision: Underpriced")
            else:
                st.write("Weighted value of ",option, " is lower than current stock price.")
                st.write("Decision: Overpriced")

    else:
        risk_free_rate = 0.029
        market_rate_of_return = 0.10
        constant = 0.21
        st.sidebar.write("US Company ")
        st.sidebar.write("Risk free rate for US: ", risk_free_rate)
        st.sidebar.write("Market rate of return for US: ", market_rate_of_return)
        st.sidebar.write("Beta: ", betas[option])
        st.sidebar.write("Number of common stocks: ", no_of_common_stock[option])
        book = load_workbook(excel_names[option])
        financial_sheets = book[sheet_names[option]]
        fs_rows = financial_sheets.rows
        st.write("Financial sheet of ",option," Company is used.")
        dcf(option,risk_free_rate,market_rate_of_return)

        sector_list = sectors[option]
        total_sector_value = 0
        no_of_sector_company = 0
        total_stock_price_of_sector = 0  #yeni
        total_evm_price = 0


        with st.expander("SEE ENTERPRISE OUTPUT HERE"):
            for company in sector_list:
                enterprise_formul(company)
                total_sector_value += enterprise_formul.valueofcompany
                no_of_sector_company += 1
                current_price = si.get_live_price(tickers[company]) #yeni
                total_stock_price_of_sector += current_price           #yeni
                average_stock_price_of_sector = total_stock_price_of_sector /no_of_sector_company #yeni
                sector_average = total_sector_value / no_of_sector_company
                evm_stock_price = ((average_stock_price_of_sector/sector_average)*enterprise_formul.valueofcompany) #son
                total_evm_price += evm_stock_price #son
                average_evm_value = total_evm_price / no_of_sector_company #son
                st.write("Enterprise Ratio of ", company ,": ",enterprise_formul.valueofcompany)
            enterprise_formul(option) # yeni enterprise value değeri için
            evm_stock_price = ((average_stock_price_of_sector/sector_average)*enterprise_formul.valueofcompany) # yeni enterprise value değeri için
            st.write("Enterprise Value for ",book.sheetnames[0] ," is: ",evm_stock_price) # yeni enterprise value değeri için
            st.write("Average stock price of ", sectors_string[option], " is ",average_stock_price_of_sector)  #yeni
            st.write("Average Enterprise Ratio for ", sectors_string[option], " is ",sector_average)
            if(evm_stock_price > average_evm_value):
                st.write("value of ",option, " is greater than average.")
                st.write("Decision: Overpriced")
            else:
                st.write("value of ",option, " is lower than average.")
                st.write("Decision: Underpriced")
        with st.expander("SEE WEIGHTED VALUE OF ALL METHODS OUTPUT HERE"):
            weighted_value = 0.4 * (dcf.fair_value_of_company) + 0.5 * dcf.price_per_share + 0.1 * evm_stock_price ## yeni weighted için
            st.write("Weighted Value", weighted_value) ## yeni weighted için
            st.write("Current stock price ",si.get_live_price(tickers[option]))
            if(weighted_value > si.get_live_price(tickers[option])):
                st.write("Weighted value of ",option, " is greater than current stock price.")
                st.write("Decision: Underpriced")
            else:
                st.write("Weighted value of ",option, " is lower than current stock price.")
                st.write("Decision: Overerpriced")
            
logout_url = 'http://localhost/Capstone/index.php'
           
#if st.sidebar.button("Log out") :
#    webbrowser.open_new_tab(logout_url)